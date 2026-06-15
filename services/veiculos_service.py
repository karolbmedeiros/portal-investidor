import unicodedata
import re
import json
import os
from typing import Optional
from services.supabase_client import get_financeiro_client

_CLIENTES_CACHE: Optional[dict] = None

_NATUREZAS_CARROS_PATH = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "static", "data", "naturezas_carros.json")
)

def listar_naturezas_carros() -> list:
    try:
        with open(_NATUREZAS_CARROS_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def adicionar_natureza_carros(nome: str) -> dict:
    nome = nome.strip()
    if not nome:
        return {"ok": False, "erro": "Nome vazio"}
    nats = listar_naturezas_carros()
    if nome in nats:
        return {"ok": False, "erro": "Já existe"}
    nats.append(nome)
    try:
        with open(_NATUREZAS_CARROS_PATH, "w", encoding="utf-8") as f:
            json.dump(nats, f, ensure_ascii=False, indent=2)
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "erro": str(e)}

def remover_natureza_carros(nome: str) -> dict:
    nats = listar_naturezas_carros()
    if nome not in nats:
        return {"ok": False, "erro": "Não encontrada"}
    nats.remove(nome)
    try:
        with open(_NATUREZAS_CARROS_PATH, "w", encoding="utf-8") as f:
            json.dump(nats, f, ensure_ascii=False, indent=2)
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "erro": str(e)}

def dados_clientes_cons() -> dict:
    """Dict keyed by normalized plate (no dash/space, uppercase)."""
    global _CLIENTES_CACHE
    if _CLIENTES_CACHE is not None:
        return _CLIENTES_CACHE
    path = os.path.join(os.path.dirname(__file__), "..", "static", "data", "clientes_carros.json")
    try:
        with open(os.path.abspath(path), encoding="utf-8") as f:
            _CLIENTES_CACHE = json.load(f)
    except Exception:
        _CLIENTES_CACHE = {}
    return _CLIENTES_CACHE


_EMPRESA_CONTA = {
    "LUZ DIVINA EMPREENDIMENTOS LTDA":            "f02a7e50-de24-4db0-913a-cc4309b21b1a",
    "JOÃO PAULO SERVIÇOS EM CONSULTORIA LTDA":    "bcc0fb80-4960-44c8-aa68-2b3c9e2d6a06",
}

_CONVENIO_KEYWORDS = [
    "GELO E GELA",
    "JUAN E IVAN",
]

def calcular_saldo_em(empresa_nome: str, data_ate: str) -> float:
    """Soma todos os lançamentos da conta (sem filtros) até data_ate inclusive.
    Usado para obter o saldo real antes das transações carros começarem.
    """
    conta_id = _EMPRESA_CONTA.get(empresa_nome)
    if not conta_id:
        return 0.0
    try:
        sb = get_financeiro_client()
        rows = (
            sb.table("lancamentos_bancarios")
            .select("valor")
            .eq("conta_bancaria_id", conta_id)
            .lte("data_transacao", data_ate)
            .execute()
            .data or []
        )
        return round(sum(float(r.get("valor") or 0) for r in rows), 2)
    except Exception as e:
        print(f"[calcular_saldo_em] erro: {e}")
        return 0.0


def listar_lancamentos_carros(empresa_nome: str) -> list:
    conta_id = _EMPRESA_CONTA.get(empresa_nome)
    if not conta_id:
        return []
    try:
        sb = get_financeiro_client()
        rows = (
            sb.table("lancamentos_bancarios")
            .select("id,data_transacao,mes_competencia,descricao,descricao_original,valor,tipo,conciliado,observacoes")
            .eq("conta_bancaria_id", conta_id)
            .order("data_transacao", desc=True)
            .execute()
            .data or []
        )
        for r in rows:
            texto = ((r.get("descricao") or "") + (r.get("descricao_original") or "")).upper()
            r["_convenio"] = any(k in texto for k in _CONVENIO_KEYWORDS)
        return rows
    except Exception as e:
        print(f"[listar_lancamentos_carros] erro: {e}")
        return []


def classificar_lancamento_carros(lancamento_id: str, natureza: str = None, splits: list = None) -> dict:
    if not lancamento_id or (not natureza and not splits):
        return {"ok": False, "erro": "id e natureza/splits obrigatorios"}
    try:
        sb = get_financeiro_client()
        obs = json.dumps(splits, ensure_ascii=False) if splits else natureza
        sb.table("lancamentos_bancarios") \
          .update({"observacoes": obs, "conciliado": True}) \
          .eq("id", lancamento_id) \
          .execute()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "erro": str(e)}


def salvar_dados_cliente(dados: dict) -> dict:
    """Persiste as alterações de um cliente no JSON local e invalida o cache."""
    global _CLIENTES_CACHE
    path = os.path.abspath(
        os.path.join(os.path.dirname(__file__), "..", "static", "data", "clientes_carros.json")
    )
    try:
        try:
            with open(path, encoding="utf-8") as f:
                store = json.load(f)
        except Exception:
            store = {}

        placa = (dados.get("placa") or dados.get("ref_id") or "").upper().replace("-", "").replace(" ", "")
        if not placa:
            return {"ok": False, "erro": "Placa/ref_id obrigatório"}

        entrada = store.get(placa, {})
        campos = ["nome", "cpf", "telefone", "endereco", "cep", "placa",
                  "ano_modelo", "marca", "cor", "chassi", "num_motor",
                  "contrato_locacao", "contrato_comercial", "tipo_contrato",
                  "unidade", "inicio", "termino"]
        for c in campos:
            if c in dados:
                entrada[c] = dados[c]
        store[placa] = entrada

        with open(path, "w", encoding="utf-8") as f:
            json.dump(store, f, ensure_ascii=False, indent=2)

        _CLIENTES_CACHE = store
        return {"ok": True, "dados": entrada}
    except Exception as e:
        return {"ok": False, "erro": str(e)}


def upload_pdf_cliente(ref_id: str, nome_arquivo: str, conteudo: bytes,
                       mime_type: str = "application/pdf") -> dict:
    from services.supabase_client import get_service_client
    import uuid
    sb = get_service_client()
    fid = str(uuid.uuid4())
    caminho = f"clientes/{ref_id}/{fid}_{nome_arquivo}"
    try:
        sb.storage.from_("documentos-contratos").upload(caminho, conteudo, {"content-type": mime_type})
        url = sb.storage.from_("documentos-contratos").get_public_url(caminho)
        return {"ok": True, "url": url, "caminho": caminho}
    except Exception as e:
        return {"ok": False, "erro": str(e)}

_CONTAS_RECEBER_PATH  = "/Users/karol/Documents/Dashboard-Ativuz/planilhas/CONTAS-A-RECEBER.xlsx"
_CONTRATOS_LOCACAO_PATH = "/Users/karol/Documents/Dashboard-Ativuz/planilhas/Contratos de Locação.xlsx"

_UNIDADE_EMPRESA = {
    "JOÃO PAULO CONSÓRCIOS": "JOÃO PAULO SERVIÇOS EM CONSULTORIA LTDA",
    "LUZ DIVINA LTDA":       "LUZ DIVINA EMPREENDIMENTOS LTDA",
    "ATIVUZ VEÍCULOS":       "ATIVUZ VEÍCULOS",
    "AZ EMPREENDIMENTOS":    "AZ EMPREENDIMENTOS",
}


def listar_contratos_excel(unidades: list) -> list:
    """Lê Contratos de Locação.xlsx e retorna contratos ativos das unidades informadas."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(_CONTRATOS_LOCACAO_PATH, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        print(f"[listar_contratos_excel] erro: {e}")
        return []

    resultado = []
    for r in rows[5:]:
        if not r[6]:  # Cliente
            continue
        cliente = str(r[6]).strip()
        if "SEGCOMP" in cliente.upper():
            continue
        unidade = str(r[54] or "").strip()
        if unidades and unidade not in unidades:
            continue
        placa_raw = r[59] or r[61] or ""
        placa = str(placa_raw).strip().upper().replace("-", "")
        inicio = r[30]
        if hasattr(inicio, "date"):
            inicio = inicio.date()
        resultado.append({
            "cliente":       cliente,
            "placa":         placa,
            "placa_fmt":     str(placa_raw).strip().upper(),
            "modelo":        str(r[36] or "").strip(),
            "inicio":        inicio.strftime("%d/%m/%Y") if inicio else "",
            "unidade":       unidade,
            "situacao":      str(r[46] or "").strip(),
            "valor_locacao": float(r[57] or r[58] or 0),
        })
    return resultado


def contratos_por_empresa(empresa_nome: str) -> list:
    unidades = [u for u, e in _UNIDADE_EMPRESA.items() if e == empresa_nome]
    return listar_contratos_excel(unidades)


def contas_receber_carros_excel(empresa_nome: str) -> list:
    """Lê CONTAS-A-RECEBER.xlsx diretamente, filtrando pela empresa."""
    import openpyxl
    from datetime import date

    unidades = [u for u, e in _UNIDADE_EMPRESA.items() if e == empresa_nome]
    if not unidades:
        return []

    try:
        wb = openpyxl.load_workbook(_CONTAS_RECEBER_PATH, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        print(f"[contas_receber_carros_excel] erro ao abrir arquivo: {e}")
        return []

    resultado = []
    hoje = date.today()
    for r in rows[5:]:
        if not r[0]:
            continue
        unidade = str(r[17] or "").strip()
        if unidade not in unidades:
            continue
        venc = r[3]
        if hasattr(venc, "date"):
            venc = venc.date()
        venc_str = venc.isoformat() if venc else None
        dias = (venc - hoje).days if venc else None
        resultado.append({
            "numero_documento": str(r[9] or ""),
            "cliente":          str(r[11] or r[12] or ""),
            "data_vencimento":  venc_str,
            "valor":            float(r[18] or 0),
            "situacao":         str(r[13] or ""),
            "tipo_fatura":      str(r[16] or ""),
            "dias_vencimento":  dias,
            "faixa_vencimento": str(r[7] or ""),
            "unidade":          unidade,
        })
    return resultado


_EMPRESA_INFO = {
    "LUZ DIVINA EMPREENDIMENTOS LTDA": {
        "cnpj":            "48.284.349/0001-29",
        "pix":             None,
        "unidade":         "LUZ DIVINA LTDA",
        "total_investido": 408_000.00,
    },
    "JOÃO PAULO SERVIÇOS EM CONSULTORIA LTDA": {
        "cnpj":            "24.954.506/0001-06",
        "pix":             "4d9e79bf-e7f3-4298-86b2-66613980b90b",
        "unidade":         "JOÃO PAULO CONSÓRCIOS",
        "total_investido": None,
    },
}

_MODELO_EMPRESA = {
    "DOLPHIN MINI": "LUZ DIVINA EMPREENDIMENTOS LTDA",
    "POLO TRACK":   "JOÃO PAULO SERVIÇOS EM CONSULTORIA LTDA",
}

_PREFIXO_EMPRESA = {
    "TSW": ("LUZ DIVINA EMPREENDIMENTOS LTDA", "BYD Dolphin Mini"),
    "SSW": ("JOÃO PAULO SERVIÇOS EM CONSULTORIA LTDA", "VW Polo Track"),
    "STX": ("JOÃO PAULO SERVIÇOS EM CONSULTORIA LTDA", "VW Polo Track"),
}


def _norm(placa):
    return (placa or "").upper().replace("-", "").replace(" ", "")


def _slugify(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")


def listar_empresas_veiculos():
    sb = get_financeiro_client()

    try:
        all_recs = (
            sb.table("sob_adm_recebimentos")
            .select("placa,taxa_valor,data_semana,recebido")
            .execute()
            .data or []
        )
    except Exception as e:
        print(f"[veiculos_service] erro recebimentos: {e}")
        return []

    if not all_recs:
        return []

    placa_to_info = {}
    try:
        contratos = (
            sb.table("contratos_locacao")
            .select("veiculo_placa,veiculo_modelo,veiculo_marca")
            .eq("deletado", False)
            .execute()
            .data or []
        )
        for c in contratos:
            key = _norm(c["veiculo_placa"])
            placa_to_info[key] = {
                "modelo": (c.get("veiculo_modelo") or "").strip(),
                "marca":  (c.get("veiculo_marca") or "").strip(),
            }
    except Exception:
        pass

    semana_atual = max(r["data_semana"] for r in all_recs)
    empresas = {}

    for rec in all_recs:
        placa = rec["placa"]
        normed = _norm(placa)
        empresa = None
        modelo = placa

        info = placa_to_info.get(normed)
        if info:
            modelo_upper = info["modelo"].upper()
            for kw, emp in _MODELO_EMPRESA.items():
                if kw in modelo_upper:
                    empresa = emp
                    modelo = info["modelo"]
                    break
        else:
            prefix = normed[:3]
            if prefix in _PREFIXO_EMPRESA:
                empresa, modelo = _PREFIXO_EMPRESA[prefix]

        if not empresa:
            continue

        if empresa not in empresas:
            info_emp = _EMPRESA_INFO.get(empresa, {})
            empresas[empresa] = {
                "nome":            empresa,
                "slug":            _slugify(empresa),
                "cnpj":            info_emp.get("cnpj"),
                "pix":             info_emp.get("pix"),
                "total_investido": info_emp.get("total_investido"),
                "veiculos": {},
                "receita_semanal": 0.0,
            }

        if placa not in empresas[empresa]["veiculos"]:
            empresas[empresa]["veiculos"][placa] = {"placa": placa, "modelo": modelo}

        if rec.get("data_semana") == semana_atual and rec.get("recebido"):
            empresas[empresa]["receita_semanal"] += float(rec.get("taxa_valor") or 0)

    result = []
    for emp in empresas.values():
        emp["veiculos"] = list(emp["veiculos"].values())
        result.append(emp)

    return result


def buscar_empresa_veiculos(slug):
    return next((e for e in listar_empresas_veiculos() if e["slug"] == slug), None)


def recebimentos_da_empresa(empresa):
    """Returns per-vehicle payment history. empresa is the dict from buscar_empresa_veiculos."""
    sb = get_financeiro_client()
    placas = [v["placa"] for v in empresa["veiculos"]]
    if not placas:
        return {}

    try:
        rows = (
            sb.table("sob_adm_recebimentos")
            .select("*")
            .in_("placa", placas)
            .order("data_semana", desc=False)
            .execute()
            .data or []
        )
    except Exception as e:
        print(f"[veiculos_service] erro historico: {e}")
        return {}

    # Group by placa → list of weekly records
    por_placa = {}
    for row in rows:
        p = row["placa"]
        if p not in por_placa:
            por_placa[p] = []
        por_placa[p].append(row)

    return por_placa


def contas_receber_empresa(empresa_nome: str) -> list:
    """Faturas em aberto da empresa na tabela contas_receber_frota."""
    info = _EMPRESA_INFO.get(empresa_nome, {})
    unidade = info.get("unidade")
    if not unidade:
        return []
    try:
        sb = get_financeiro_client()
        res = (
            sb.table("contas_receber_frota")
            .select("numero_documento,cliente,data_vencimento,valor,situacao,tipo_fatura,dias_vencimento,faixa_vencimento")
            .eq("unidade", unidade)
            .order("data_vencimento")
            .execute()
        )
        return res.data or []
    except Exception as e:
        print(f"[contas_receber_empresa] erro: {e}")
        return []
